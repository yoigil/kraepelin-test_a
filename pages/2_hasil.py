import streamlit as st
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference
from zipencrypt import ZipFile

st.set_page_config(page_title="Tes Selesai", layout="centered")
st.markdown("<style>section[data-testid='stSidebar'] {display: none;}</style>", unsafe_allow_html=True)

TOTAL_COLUMNS = 50

def generate_zipped_xlsx():
    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "Data Per Kolom"
    ws_data.views.sheetView[0].showGridLines = True
    ws_data.append(["Kolom", "Jumlah Benar", "Jumlah Salah", "Total Jumlah"])
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for c_num in range(1, 5):
        cell = ws_data.cell(row=1, column=c_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    for i in range(TOTAL_COLUMNS):
        ws_data.append([i+1, st.session_state["corrects"][i], st.session_state["errors"][i], st.session_state["attempts"][i]])
        
    ws_sum = wb.create_sheet(title="Ringkasan Hasil", index=0)
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum["A1"] = "Laporan Hasil Tes Kraepelin"
    ws_sum["A1"].font = Font(name="Arial", size=16, bold=True, color="1F497D")
    ws_sum["A2"] = f"Nama Peserta: {st.session_state.get('user_name', '')}"
    ws_sum["A2"].font = Font(name="Arial", size=11, bold=True)
    ws_sum["A3"] = f"NIK Peserta: {st.session_state.get('user_nik', '')}"
    ws_sum["A3"].font = Font(name="Arial", size=11, bold=True)
    
    headers = ["Points", "Score", "Criteria"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=5, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    metrics = [
        (
            "PANKER (Kecepatan)", 
            "=SUM('Data Per Kolom'!D2:D51)/MAX('Data Per Kolom'!A2:A51)", 
            '=IF(B6<7, "Rendah", IF(B6>10.9, "Tinggi", "Sedang"))'
        ),
        (
            "TIANKER (Ketelitian)", 
            "=SUM('Data Per Kolom'!C2:C51)", 
            '=IF(B7>=36, "Rendah", IF(B7<=10, "Tinggi", "Sedang"))'
        ),
        (
            "JANKER (Keajegan)", 
            "=MAX('Data Per Kolom'!D2:D51)-MIN('Data Per Kolom'!D2:D51)", 
            '=IF(B8<9, "Tinggi", IF(B8>12, "Rendah", "Sedang"))'
        ),
        (
            "HANKER (Ketahanan)", 
            "=50*SLOPE('Data Per Kolom'!D2:D51,'Data Per Kolom'!A2:A51)", 
            '=IF(B9<-1.947, "Rendah", IF(B9>0.223, "Tinggi", "Sedang"))'
        )
    ]
    
    for idx, (label, score_formula, criteria_formula) in enumerate(metrics, start=6):
        ws_sum.cell(row=idx, column=1, value=label).font = Font(name="Arial", bold=True)
        ws_sum.cell(row=idx, column=2, value=score_formula).alignment = center_align
        ws_sum.cell(row=idx, column=3, value=criteria_formula).alignment = center_align
        
    chart = LineChart()
    chart.title = "Grafik Ketahanan Kerja (Hanker)"
    chart.width = 15; chart.height = 8; chart.legend = None
    chart.add_data(Reference(ws_data, min_col=4, min_row=1, max_row=51), titles_from_data=True)
    chart.set_categories(Reference(ws_data, min_col=1, min_row=2, max_row=51))
    ws_sum.add_chart(chart, "A12")
    
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 15
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    
    zip_buffer = io.BytesIO()
    clean_name = st.session_state.get('user_name', 'peserta').replace(' ', '_')
    internal_excel_name = f"Laporan_Kraepelin_{clean_name}.xlsx"
    password = b"HR3918"
    
    with ZipFile(zip_buffer, mode="w") as zf:
        zf.writestr(internal_excel_name, excel_buffer.getvalue(), pwd=password)
        
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

st.balloons()
st.title("Tes Selesai!")
st.success(f"Selamat {st.session_state.get('user_name', '')}, Anda telah menyelesaikan seluruh rangkaian ujian Kraepelin!")

excel_data = generate_zipped_xlsx()
clean_nik = str(st.session_state.get('user_nik', 'peserta')).strip().replace(' ', '_')

st.download_button(
    label="Download ZIP Hasil Evaluasi",
    data=excel_data,
    file_name=f"Laporan_Kraepelin_{clean_nik}.zip",
    mime="application/zip",
    type="primary"
)

if st.button("Ulangi Tes Baru"):
    st.session_state.clear()
    st.switch_page("app.py")
